import os
from pathlib import Path
import openpyxl
import xlrd

from .validator import DateValidator
from .analyzer import FileNumberAnalyzer

class ExcelProcessor:
    SAYFA_ADI = "DOSYA MUHTEVİYATI DÖKÜM FORMU"
    BASLIK_TARIH = "TARİH"
    BASLIK_SAYI = "SAYI"
    BASLIK_ACIKLAMA = "AÇIKLAMALAR"
    BASLIK_SIRA = "SIRA"

    @classmethod
    def get_matrix(cls, dosya_yolu):
        uzanti = Path(dosya_yolu).suffix.lower()
        cells = []
        
        if uzanti == '.xls':
            wb = xlrd.open_workbook(dosya_yolu)
            target_sheet = None
            for s in wb.sheet_names():
                if s.strip().upper() == cls.SAYFA_ADI.upper():
                    target_sheet = s
                    break
            if not target_sheet:
                return None
            
            ws = wb.sheet_by_name(target_sheet)
            for r in range(ws.nrows):
                row_data = []
                for c in range(ws.ncols):
                    cell_type = ws.cell_type(r, c)
                    val = ws.cell_value(r, c)
                    if cell_type == xlrd.XL_CELL_DATE:
                        try:
                            val = xlrd.xldate.xldate_as_datetime(val, wb.datemode)
                        except (ValueError, TypeError, OverflowError):
                            pass
                    row_data.append(val)
                cells.append(row_data)
                
        else: # .xlsx, .xlsm
            wb = openpyxl.load_workbook(dosya_yolu, data_only=True)
            target_sheet = None
            for s in wb.sheetnames:
                if s.strip().upper() == cls.SAYFA_ADI.upper():
                    target_sheet = s
                    break
            
            if not target_sheet:
                wb.close()
                return None
                
            ws = wb[target_sheet]
            for row in ws.iter_rows(values_only=True):
                cells.append(list(row))
            wb.close()
            
        return cells

    @classmethod
    def process_file(cls, dosya_yolu):
        uzanti = Path(dosya_yolu).suffix.lower()
        if uzanti not in (".xls", ".xlsx", ".xlsm"):
            return {"status": "error", "message": f"Desteklenmeyen format: {uzanti}. Sadece .xls, .xlsx, .xlsm desteklenir."}

        try:
            matrix = cls.get_matrix(dosya_yolu)
        except Exception as e:
            return {"status": "error", "message": f"Dosya açılamadı: {e}"}
            
        if matrix is None:
            return {"status": "error", "message": f"'{cls.SAYFA_ADI}' sayfası bulunamadı!"}

        # Bul başlıkları
        tarih_sutun = sayi_sutun = acik_sutun = sira_sutun = None
        baslik_satiri = -1
        
        for r_idx, row in enumerate(matrix):
            if r_idx > 50: break # İlk 50 satırda başlık aranır
            for c_idx, val in enumerate(row):
                if val:
                    hucre_metni = str(val).strip().upper()
                    if hucre_metni == cls.BASLIK_TARIH.upper():
                        tarih_sutun = c_idx
                        if baslik_satiri == -1: baslik_satiri = r_idx
                    elif hucre_metni == cls.BASLIK_SAYI.upper():
                        sayi_sutun = c_idx
                        if baslik_satiri == -1: baslik_satiri = r_idx
                    elif hucre_metni == cls.BASLIK_ACIKLAMA.upper():
                        acik_sutun = c_idx
                        if baslik_satiri == -1: baslik_satiri = r_idx
                    elif hucre_metni == cls.BASLIK_SIRA.upper():
                        sira_sutun = c_idx
                        if baslik_satiri == -1: baslik_satiri = r_idx
                        
        if baslik_satiri == -1:
            return {"status": "error", "message": "Hiçbir sütun başlığı bulunamadı!"}
            
        veri_baslangic = baslik_satiri + 1

        def form_sirasi_oku(row_data):
            if sira_sutun is None or sira_sutun >= len(row_data): return ""
            val = row_data[sira_sutun]
            if val is None or str(val).strip() == "": return ""
            if isinstance(val, float) and val == int(val): return str(int(val))
            return str(val).strip()

        def numara_yap(sutun):
            from string import ascii_uppercase
            result = ""
            sutun += 1 # 0-indexed to 1-indexed
            while sutun > 0:
                sutun, remainder = divmod(sutun - 1, 26)
                result = ascii_uppercase[remainder] + result
            return result

        tarih_hatalari = []
        if tarih_sutun is not None:
            for r_idx in range(veri_baslangic, len(matrix)):
                row_data = matrix[r_idx]
                if tarih_sutun >= len(row_data): continue
                
                deger = row_data[tarih_sutun]
                if deger is None or str(deger).strip() == "": continue
                
                gecerli, hata_turu, hata_detayi = DateValidator.is_valid(deger)
                if not gecerli:
                    dosya_no = row_data[sayi_sutun] if sayi_sutun is not None and sayi_sutun < len(row_data) else None
                    if isinstance(dosya_no, float) and dosya_no == int(dosya_no):
                        dosya_no = str(int(dosya_no))
                    elif dosya_no is not None:
                        dosya_no = str(dosya_no).strip()

                    hucre_adresi = f"{numara_yap(tarih_sutun)}{r_idx+1}"
                    tarih_hatalari.append({
                        "id": f"T{r_idx+1}",
                        "excel_satir": r_idx + 1,
                        "form_sirasi": form_sirasi_oku(row_data),
                        "dosya_no": dosya_no or "-",
                        "hucre": hucre_adresi,
                        "hata_turu": hata_turu,
                        "hata_detayi": hata_detayi,
                        "mevcut_deger": str(deger)
                    })

        numara_hatalari = []
        if sayi_sutun is not None and acik_sutun is not None:
            veriler = []
            for r_idx in range(veri_baslangic, len(matrix)):
                row_data = matrix[r_idx]
                
                dosya_no = row_data[sayi_sutun] if sayi_sutun < len(row_data) else None
                if isinstance(dosya_no, float) and dosya_no == int(dosya_no):
                    dosya_no = str(int(dosya_no))
                elif dosya_no is not None:
                    dosya_no = str(dosya_no).strip()

                isim = row_data[acik_sutun] if acik_sutun < len(row_data) else None
                if isim is not None:
                    isim = str(isim).strip()
                
                veriler.append((r_idx + 1, form_sirasi_oku(row_data), dosya_no, isim))

            numara_hatalari = FileNumberAnalyzer.check_file_numbers(veriler)
            for idx, h in enumerate(numara_hatalari):
                 h["id"] = f"N{idx}"

        toplam_hata = len(tarih_hatalari) + len(numara_hatalari)

        return {
            "status": "success",
            "tarih_hatalari": tarih_hatalari,
            "numara_hatalari": numara_hatalari,
            "toplam_hata": toplam_hata
        }
